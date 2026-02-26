#!/usr/bin/env python3
"""
メルカリ出品テキスト自動生成ツール

配布用リスト（xlsx）の商品データから、メルカリ向けのタイトルと説明文を
OpenAI GPT-5 mini で自動生成し、xlsxに書き戻す。

使い方:
  python generate.py input.xlsx
  python generate.py input.xlsx --rows 3-5
  python generate.py input.xlsx --force
  python generate.py input.xlsx --dry-run
"""

import os
import sys
import json
import re
import time
import shutil
import argparse
from pathlib import Path

from openpyxl import load_workbook
from openai import OpenAI

from prompt import SYSTEM_PROMPT, TITLE_SHORTEN_PROMPT, build_user_message

# ─── 設定 ───────────────────────────────────────────────
MODEL = "gpt-5-mini"
MAX_TOKENS = 2048
TEMPERATURE = 0.4
RETRY_MAX = 3
SLEEP_SEC = 1.0
TITLE_MAX_LEN = 40

# 列マッピング（1-indexed、仕様書準拠）
COL_CONFIRM = 1       # A: 確認
COL_BOX_ID = 2        # B: 箱ID
COL_MANAGE_NO = 3     # C: 管理番号
COL_BRAND = 4         # D: ブランド
COL_AI_TITLE = 5      # E: AIタイトル候補 → 生成タイトルで上書き
COL_ITEM = 6          # F: アイテム
COL_SIZE = 7          # G: サイズ
COL_CONDITION = 8     # H: 状態
COL_DAMAGE = 9        # I: 傷汚れ詳細
COL_MEASUREMENTS = 10 # J: 採寸情報
COL_DESCRIPTION = 11  # K: 即出品用説明文 → 生成説明文で上書き
COL_PRICE = 12        # L: 金額

# ヘッダー行の設定
HEADER_ROW = 2         # 2行目が列ヘッダー
DATA_START_ROW = 3     # 3行目からデータ

# スキップ判定文字列
SKIP_PREFIX = "ご覧いただきありがとうございます"

# ─── APIクライアント ────────────────────────────────────
api_key = os.environ.get("OPENAI_API_KEY", "")
client = None


def init_client():
    global client, api_key
    if not api_key:
        print("ERROR: OPENAI_API_KEY が未設定です")
        print("  export OPENAI_API_KEY='sk-xxxxx'")
        sys.exit(1)
    client = OpenAI(api_key=api_key)


def cell_str(ws, row, col):
    """セルの値を文字列として取得"""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def read_row_data(ws, row):
    """1行分のデータを辞書で取得"""
    return {
        "manage_no": cell_str(ws, row, COL_MANAGE_NO),
        "brand": cell_str(ws, row, COL_BRAND),
        "ai_keywords": cell_str(ws, row, COL_AI_TITLE),
        "item": cell_str(ws, row, COL_ITEM),
        "size": cell_str(ws, row, COL_SIZE),
        "condition": cell_str(ws, row, COL_CONDITION),
        "damage_detail": cell_str(ws, row, COL_DAMAGE),
        "measurements": cell_str(ws, row, COL_MEASUREMENTS),
        "price": cell_str(ws, row, COL_PRICE),
    }


def should_skip(ws, row, force=False):
    """スキップ条件チェック"""
    if force:
        return False
    desc = cell_str(ws, row, COL_DESCRIPTION)
    if desc.startswith(SKIP_PREFIX):
        return True
    return False


def is_empty_row(ws, row):
    """データのない空行か判定"""
    brand = cell_str(ws, row, COL_BRAND)
    item = cell_str(ws, row, COL_ITEM)
    return not brand and not item


def call_api(user_message, retry=0):
    """OpenAI API呼び出し（リトライ付き）"""
    try:
        res = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_message},
            ],
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            response_format={"type": "json_object"},
        )
        content = res.choices[0].message.content.strip()
        return json.loads(content)
    except json.JSONDecodeError as e:
        if retry < RETRY_MAX:
            wait = 2 ** retry
            print(f"    JSON解析失敗、{wait}秒後にリトライ ({retry + 1}/{RETRY_MAX})")
            time.sleep(wait)
            return call_api(user_message, retry + 1)
        print(f"    JSON解析失敗（リトライ上限）: {e}")
        return None
    except Exception as e:
        if retry < RETRY_MAX:
            wait = 2 ** retry
            print(f"    APIエラー、{wait}秒後にリトライ ({retry + 1}/{RETRY_MAX}): {e}")
            time.sleep(wait)
            return call_api(user_message, retry + 1)
        print(f"    APIエラー（リトライ上限）: {e}")
        return None


def shorten_title(title):
    """タイトルが40文字超の場合、APIで短縮"""
    try:
        res = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "user", "content": TITLE_SHORTEN_PROMPT.format(title=title)},
            ],
            max_tokens=100,
            temperature=0.2,
        )
        shortened = res.choices[0].message.content.strip()
        # 引用符があれば除去
        shortened = shortened.strip('"\'「」')
        if len(shortened) <= TITLE_MAX_LEN and len(shortened) > 0:
            return shortened
        # それでも超えていたら機械的にカット
        return title[:TITLE_MAX_LEN]
    except Exception as e:
        print(f"    タイトル短縮失敗: {e}")
        return title[:TITLE_MAX_LEN]


def parse_row_range(range_str, max_row):
    """'3-5' → [3,4,5]、'3' → [3]"""
    parts = range_str.split("-")
    if len(parts) == 2:
        start = max(int(parts[0]), DATA_START_ROW)
        end = min(int(parts[1]), max_row)
        return list(range(start, end + 1))
    elif len(parts) == 1:
        r = int(parts[0])
        if DATA_START_ROW <= r <= max_row:
            return [r]
    return []


def main():
    parser = argparse.ArgumentParser(description="メルカリ出品テキスト自動生成ツール")
    parser.add_argument("input_file", help="入力xlsxファイル")
    parser.add_argument("--rows", help="処理する行範囲（例: 3-5）", default=None)
    parser.add_argument("--force", action="store_true", help="既存の説明文があっても再生成")
    parser.add_argument("--dry-run", action="store_true", help="APIを呼ばずプロンプトだけ表示")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"ERROR: ファイルが見つかりません: {input_path}")
        sys.exit(1)

    # バックアップ
    backup_path = input_path.with_name(input_path.stem + "_backup" + input_path.suffix)
    shutil.copy2(input_path, backup_path)
    print(f"バックアップ: {backup_path}")

    # 出力ファイル名
    output_path = input_path.with_name(input_path.stem + "_generated" + input_path.suffix)

    # ワークブックを開く
    wb = load_workbook(str(input_path))
    ws = wb.active

    max_row = ws.max_row
    print(f"データ行: {DATA_START_ROW}〜{max_row}行目（{max_row - DATA_START_ROW + 1}件）")

    # 処理対象行の決定
    if args.rows:
        target_rows = parse_row_range(args.rows, max_row)
        print(f"指定行のみ処理: {target_rows}")
    else:
        target_rows = list(range(DATA_START_ROW, max_row + 1))

    if not args.dry_run:
        init_client()

    processed = 0
    skipped = 0
    errors = 0
    total = len(target_rows)

    for i, row in enumerate(target_rows):
        # 空行スキップ
        if is_empty_row(ws, row):
            continue

        manage_no = cell_str(ws, row, COL_MANAGE_NO)
        brand = cell_str(ws, row, COL_BRAND)
        label = f"{manage_no} ({brand})" if manage_no else f"行{row}"

        # スキップ判定
        if should_skip(ws, row, args.force):
            skipped += 1
            continue

        row_data = read_row_data(ws, row)
        user_msg = build_user_message(row_data)

        print(f"\n[{processed + skipped + 1}/{total}] {label}")

        if args.dry_run:
            print("  --- プロンプト ---")
            print(f"  ブランド: {row_data['brand']}")
            print(f"  AIキーワード: {row_data['ai_keywords']}")
            print(f"  アイテム: {row_data['item']}")
            print(f"  サイズ: {row_data['size']}")
            print(f"  状態: {row_data['condition']}")
            print(f"  傷汚れ: {row_data['damage_detail']}")
            print(f"  採寸: {row_data['measurements']}")
            print(f"  価格: {row_data['price']}")
            processed += 1
            continue

        # API呼び出し
        result = call_api(user_msg)
        if not result:
            print(f"  -> 生成失敗（スキップ）")
            errors += 1
            continue

        title = result.get("title", "")
        description = result.get("description", "")

        if not title or not description:
            print(f"  -> 不完全な応答（スキップ）")
            errors += 1
            continue

        # タイトル文字数チェック
        if len(title) > TITLE_MAX_LEN:
            print(f"  タイトル {len(title)}文字 → 短縮中...")
            title = shorten_title(title)
            time.sleep(SLEEP_SEC)

        # 改行コード処理（\nを実際の改行に）
        description = description.replace("\\n", "\n")

        # 書き込み
        ws.cell(row=row, column=COL_AI_TITLE, value=title)
        ws.cell(row=row, column=COL_DESCRIPTION, value=description)

        print(f"  タイトル ({len(title)}文字): {title}")
        print(f"  説明文: {len(description)}文字")
        processed += 1

        time.sleep(SLEEP_SEC)

    # 保存
    if not args.dry_run and processed > 0:
        wb.save(str(output_path))
        print(f"\n{'=' * 60}")
        print(f"完了: {processed}件 処理 / {skipped}件 スキップ / {errors}件 エラー")
        print(f"出力: {output_path}")
    elif args.dry_run:
        print(f"\n{'=' * 60}")
        print(f"ドライラン完了: {processed}件 確認 / {skipped}件 スキップ")
    else:
        print(f"\n処理対象がありませんでした（{skipped}件 スキップ / {errors}件 エラー）")


if __name__ == "__main__":
    main()
